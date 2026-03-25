"""
Parse 'Payments - Raport Mujor.xlsx' workbook into clean JSON modules.
Reads all 5 data sheets and produces payments_data.json with:
  accounts, cards, terminals, terminal_tx, payments_count, payments_value
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')

import json
import openpyxl
from pathlib import Path

WORKBOOK = Path(__file__).parent / "Payments - Raport Mujor.xlsx"
OUTPUT   = Path(__file__).parent / "payments_data.json"

# Albanian month name -> number
MONTH_MAP = {
    'janar': 1, 'shkurt': 2, 'mars': 3, 'prill': 4, 'maj': 5,
    'qershor': 6, 'korrik': 7, 'gusht': 8, 'shtator': 9,
    'tetor': 10, 'nëntor': 11, 'nentor': 11, 'dhjetor': 12,
}

MONTH_EN = {
    1: 'January', 2: 'February', 3: 'March', 4: 'April',
    5: 'May', 6: 'June', 7: 'July', 8: 'August',
    9: 'September', 10: 'October', 11: 'November', 12: 'December',
}


def cell_val(ws, row, col):
    """Get cell value, return None if empty."""
    v = ws.cell(row=row, column=col).value
    return v


def num(v):
    """Convert to number, return None if not numeric."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    if isinstance(v, str):
        v = v.strip().replace(',', '')
        if v == '' or v == '-':
            return None
        try:
            return float(v)
        except ValueError:
            return None
    return None


def parse_period_rows(ws, start_row, end_row, year_col=1, month_col=2):
    """
    Parse rows that have year in col 1 (only on first row of year block)
    and month name in col 2. Returns list of (row_num, period_string) tuples.
    """
    rows = []
    current_year = None
    for r in range(start_row, end_row + 1):
        yr = cell_val(ws, r, year_col)
        mo = cell_val(ws, r, month_col)

        if yr is not None:
            # Handle cases like '2025*', '*', or plain 2025
            yr_str = str(yr).strip().replace('*', '').strip()
            if yr_str:
                yr_val = num(yr_str)
                if yr_val is not None:
                    current_year = int(yr_val)

        if mo is None:
            continue

        mo_str = str(mo).strip().lstrip('*').strip().lower()
        if mo_str not in MONTH_MAP:
            continue

        mo_num = MONTH_MAP[mo_str]
        period = f"{current_year} {MONTH_EN[mo_num]}"
        rows.append((r, period))

    return rows


def find_data_rows(ws, start_row=12, max_row=60):
    """Find rows with data by scanning for year/month patterns."""
    return parse_period_rows(ws, start_row, max_row)


def parse_accounts(wb):
    """Parse Llogaritë e klientëve sheet."""
    ws = wb['Llogaritë e klientëve']
    rows = parse_period_rows(ws, 12, 30)
    data = []
    for r, period in rows:
        data.append({
            'period': period,
            'total': num(cell_val(ws, r, 3)),
            'individual': num(cell_val(ws, r, 4)),
            'business': num(cell_val(ws, r, 5)),
            'current': num(cell_val(ws, r, 6)),
            'current_ind': num(cell_val(ws, r, 7)),
            'current_bus': num(cell_val(ws, r, 8)),
            'ebanking': num(cell_val(ws, r, 9)),
            'ebanking_ind': num(cell_val(ws, r, 10)),
            'ebanking_bus': num(cell_val(ws, r, 11)),
            'ebanking_pct': num(cell_val(ws, r, 12)),
            'basic': num(cell_val(ws, r, 15)),
            'savings': num(cell_val(ws, r, 16)),
            'savings_ind': num(cell_val(ws, r, 17)),
            'savings_bus': num(cell_val(ws, r, 18)),
            'term': num(cell_val(ws, r, 19)),
            'term_ind': num(cell_val(ws, r, 20)),
            'term_bus': num(cell_val(ws, r, 21)),
            'emoney': num(cell_val(ws, r, 22)),
        })
    return data


def parse_cards(wb):
    """Parse Kartelat bankare sheet."""
    ws = wb['Kartelat bankare']
    rows = parse_period_rows(ws, 12, 30)
    data = []
    for r, period in rows:
        # Col 13 = debit and/or delayed debit, Col 14 = credit and/or delayed debit
        data.append({
            'period': period,
            'total': num(cell_val(ws, r, 3)),
            'cash_fn': num(cell_val(ws, r, 4)),
            'payment_fn': num(cell_val(ws, r, 5)),
            'local': num(cell_val(ws, r, 6)),
            'visa': num(cell_val(ws, r, 7)),
            'mastercard': num(cell_val(ws, r, 8)),
            'other': num(cell_val(ws, r, 9)),
            'debit': num(cell_val(ws, r, 10)),
            'credit': num(cell_val(ws, r, 11)),
            'delayed_debit': num(cell_val(ws, r, 12)),
            'contact': num(cell_val(ws, r, 15)),
            'contactless': num(cell_val(ws, r, 16)),
            'emoney_fn': num(cell_val(ws, r, 17)),
        })
    return data


def parse_terminals(wb):
    """Parse Terminalet për pagesa sheet."""
    ws = wb['Terminalet për pagesa']
    rows = parse_period_rows(ws, 12, 30)
    data = []
    for r, period in rows:
        data.append({
            'period': period,
            'total_atm': num(cell_val(ws, r, 3)),
            'atm_cash': num(cell_val(ws, r, 4)),
            'atm_transfer': num(cell_val(ws, r, 5)),
            'atm_deposit': num(cell_val(ws, r, 6)),
            'total_pos': num(cell_val(ws, r, 7)),
            'pos_cash': num(cell_val(ws, r, 8)),
            'eftpos': num(cell_val(ws, r, 9)),
            'pos_contact': num(cell_val(ws, r, 10)),
            'pos_contactless': num(cell_val(ws, r, 11)),
            'virtual_pos': num(cell_val(ws, r, 12)),
            'total_emoney_term': num(cell_val(ws, r, 13)),
            'emoney_load': num(cell_val(ws, r, 14)),
            'emoney_pay': num(cell_val(ws, r, 15)),
            'merchants_physical': num(cell_val(ws, r, 16)),
            'merchants_virtual': num(cell_val(ws, r, 17)),
        })
    return data


def parse_terminal_tx(wb):
    """Parse Trans. sipas terminaleve sheet."""
    ws = wb['Trans. sipas terminaleve']
    rows = parse_period_rows(ws, 12, 30)
    data = []
    for r, period in rows:
        data.append({
            'period': period,
            'total_count': num(cell_val(ws, r, 3)),
            'total_value': num(cell_val(ws, r, 4)),
            'atm_cash_count': num(cell_val(ws, r, 5)),
            'atm_cash_value': num(cell_val(ws, r, 6)),
            'atm_deposit_count': num(cell_val(ws, r, 7)),
            'atm_deposit_value': num(cell_val(ws, r, 8)),
            'atm_transfer_count': num(cell_val(ws, r, 9)),
            'atm_transfer_value': num(cell_val(ws, r, 10)),
            'pos_cash_count': num(cell_val(ws, r, 11)),
            'pos_cash_value': num(cell_val(ws, r, 12)),
            'pos_card_count': num(cell_val(ws, r, 13)),
            'pos_card_value': num(cell_val(ws, r, 14)),
            'emoney_load_count': num(cell_val(ws, r, 15)),
            'emoney_load_value': num(cell_val(ws, r, 16)),
            'emoney_pay_count': num(cell_val(ws, r, 17)),
            'emoney_pay_value': num(cell_val(ws, r, 18)),
        })
    return data


def parse_payments_by_instrument(wb):
    """Parse Pagesat sipas instrumenteve - two stacked tables (count rows 12-26, value rows 34-47)."""
    ws = wb['Pagesat sipas instrumenteve']

    # COUNT table
    count_rows = parse_period_rows(ws, 12, 26)
    counts = []
    for r, period in count_rows:
        counts.append({
            'period': period,
            'total': num(cell_val(ws, r, 3)),
            'credit_transfer': num(cell_val(ws, r, 6)),
            'paper_ct': num(cell_val(ws, r, 7)),
            'electronic_ct': num(cell_val(ws, r, 10)),
            'card_total': num(cell_val(ws, r, 13)),
            'card_debit': num(cell_val(ws, r, 16)),
            'card_credit': num(cell_val(ws, r, 19)),
            'card_delayed': num(cell_val(ws, r, 22)),
            'physical_card': num(cell_val(ws, r, 31)),
            'remote_card': num(cell_val(ws, r, 32)),
            'intl_incoming': num(cell_val(ws, r, 33)),
            'intl_outgoing': num(cell_val(ws, r, 36)),
            'emoney': num(cell_val(ws, r, 39)),
            'ecommerce': num(cell_val(ws, r, 40)),
            'digital_wallet': num(cell_val(ws, r, 41)),
            'other': num(cell_val(ws, r, 42)),
        })

    # VALUE table (header at row 33, data rows 34-47)
    value_rows = parse_period_rows(ws, 34, 50)
    values = []
    for r, period in value_rows:
        values.append({
            'period': period,
            'total': num(cell_val(ws, r, 3)),
            'credit_transfer': num(cell_val(ws, r, 6)),
            'paper_ct': num(cell_val(ws, r, 7)),
            'electronic_ct': num(cell_val(ws, r, 10)),
            'card_total': num(cell_val(ws, r, 13)),
            'card_debit': num(cell_val(ws, r, 16)),
            'card_credit': num(cell_val(ws, r, 19)),
            'card_delayed': num(cell_val(ws, r, 22)),
            'physical_card': num(cell_val(ws, r, 31)),
            'remote_card': num(cell_val(ws, r, 32)),
            'intl_incoming': num(cell_val(ws, r, 33)),
            'intl_outgoing': num(cell_val(ws, r, 36)),
            'emoney': num(cell_val(ws, r, 39)),
            'ecommerce': num(cell_val(ws, r, 40)),
            'digital_wallet': num(cell_val(ws, r, 41)),
            'other': num(cell_val(ws, r, 42)),
        })

    return counts, values


def main():
    print(f"Loading workbook: {WORKBOOK}")
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)

    print(f"Sheet names: {wb.sheetnames}")

    # Parse all sheets
    accounts = parse_accounts(wb)
    print(f"\n=== Accounts: {len(accounts)} rows ===")
    for a in accounts:
        print(f"  {a['period']}: total={a['total']:,.0f}, ebanking={a['ebanking']}, emoney={a['emoney']}" if a['total'] else f"  {a['period']}: no data")

    cards = parse_cards(wb)
    print(f"\n=== Cards: {len(cards)} rows ===")
    for c in cards:
        print(f"  {c['period']}: total={c['total']:,.0f}, debit={c['debit']}, credit={c['credit']}" if c['total'] else f"  {c['period']}: no data")

    terminals = parse_terminals(wb)
    print(f"\n=== Terminals: {len(terminals)} rows ===")
    for t in terminals:
        print(f"  {t['period']}: ATMs={t['total_atm']}, POS={t['total_pos']}, virtual={t['virtual_pos']}" if t['total_atm'] else f"  {t['period']}: no data")

    terminal_tx = parse_terminal_tx(wb)
    print(f"\n=== Terminal Transactions: {len(terminal_tx)} rows ===")
    for t in terminal_tx:
        tv = t['total_value']
        tc = t['total_count']
        if tc:
            print(f"  {t['period']}: count={tc:,.0f}, value={tv:,.0f}" if tv else f"  {t['period']}: count={tc:,.0f}")
        else:
            print(f"  {t['period']}: no data")

    payments_count, payments_value = parse_payments_by_instrument(wb)
    print(f"\n=== Payments Count: {len(payments_count)} rows ===")
    for p in payments_count:
        print(f"  {p['period']}: total={p['total']:,.0f}, card={p['card_total']}, ecommerce={p['ecommerce']}" if p['total'] else f"  {p['period']}: no data")

    print(f"\n=== Payments Value: {len(payments_value)} rows ===")
    for p in payments_value:
        print(f"  {p['period']}: total={p['total']:,.0f}" if p['total'] else f"  {p['period']}: no data")

    # Build output
    result = {
        'accounts': accounts,
        'cards': cards,
        'terminals': terminals,
        'terminal_tx': terminal_tx,
        'payments_count': payments_count,
        'payments_value': payments_value,
    }

    with open(OUTPUT, 'w', encoding='utf-8') as f:
        json.dump(result, f, indent=2, ensure_ascii=False)

    print(f"\n✓ Saved to {OUTPUT}")
    print(f"  Total modules: {len(result)}")
    for k, v in result.items():
        print(f"  {k}: {len(v)} records")


if __name__ == '__main__':
    main()
